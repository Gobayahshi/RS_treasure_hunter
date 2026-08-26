"""엑셀 마스터 데이터 파싱 및 upsert.

지원하는 입력 형태
- 시트 3개짜리 엑셀 1개 (시트명: 대리점 / 영업사원 / 판매점)
- 파일 3개 (파일명에 대리점, 영업사원, 판매점이 들어가면 자동 분류)

컬럼명은 공백/대소문자를 무시하고 아래 별칭을 모두 받는다.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from werkzeug.security import generate_password_hash

DEALER_CODE_ALIASES = {"대리점id", "대리점코드", "대리점아이디", "소속대리점id", "소속대리점코드", "dealercode", "dealerid", "dealer_code", "dealer_id"}
DEALER_NAME_ALIASES = {"대리점명", "대리점이름", "소속대리점명", "소속대리점", "dealername", "dealer_name"}
REP_CODE_ALIASES = {"고유id", "사원고유id", "사원id", "사원코드", "사번", "아이디", "id", "employeecode", "employee_code", "employeeid"}
REP_NAME_ALIASES = {"이름", "성명", "사원명", "영업사원명", "name"}
STORE_CODE_ALIASES = {"판매점코드", "매장코드", "점포코드", "storecode", "store_code"}
STORE_NAME_ALIASES = {"판매점명", "매장명", "점포명", "storename", "store_name"}
STORE_ADDR_ALIASES = {"기본주소", "주소", "판매점주소", "매장주소", "address"}
DETAIL_ADDR_ALIASES = {"상세주소", "층호수", "detailaddress", "detail_address"}
LAT_ALIASES = {"위도", "lat", "latitude"}
LNG_ALIASES = {"경도", "lng", "lon", "longitude"}


def cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.startswith("'"):
        text = text[1:].strip()
    return text


def normalize_header(value: Any) -> str:
    raw = (
        cell_str(value)
        .lower()
        .replace(" ", "")
        .replace("_", "")
        .replace("-", "")
        .replace("\n", "")
        .replace("\r", "")
    )
    return raw


def _first_header_row(ws: Worksheet) -> tuple[int, dict[str, int]]:
    """첫 비어있지 않은 행을 헤더로 보고 {정규화헤더: 컬럼인덱스}를 반환한다."""
    for row in ws.iter_rows(min_row=1, max_row=20, values_only=False):
        values = [cell_str(c.value) for c in row]
        if any(values):
            mapping = {}
            for cell in row:
                key = normalize_header(cell.value)
                if key:
                    mapping[key] = cell.column
            return row[0].row, mapping
    return 1, {}


def sheet_to_rows(ws: Worksheet) -> list[dict[str, str]]:
    header_row, headers = _first_header_row(ws)
    if not headers:
        return []
    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
        by_col = {c.column: c.value for c in row}
        raw = {h: cell_str(by_col.get(col, "")) for h, col in headers.items()}
        if not any(raw.values()):
            continue
        rows.append(raw)
    return rows


def pick(row: dict[str, str], aliases: set[str]) -> str:
    for key, value in row.items():
        if key in aliases and value:
            return value
    return ""


def guess_kind(filename: str, sheet_name: str, headers: set[str]) -> str | None:
    sheet = sheet_name.lower().replace(" ", "").replace("\n", "")
    file_name = filename.lower().replace(" ", "")
    if "안내" in sheet or "guide" in sheet or "readme" in sheet:
        return None
    if "재고" in sheet or "재고" in filename.lower().replace(" ", ""):
        return "inventory"
    if "보유처매장코드" in headers or "대표상품명" in headers:
        return "inventory"
    # 시트명이 파일명보다 우선이다. (파일명에 '판매점'이 들어 있어도 대리점 시트를 판매점으로 오인하지 않게)
    if "영업" in sheet or "사원" in sheet or "rep" in sheet:
        return "reps"
    if "판매" in sheet or "매장" in sheet or "store" in sheet or sheet.endswith("주소"):
        return "stores"
    if "대리점" in sheet or "dealer" in sheet:
        return "dealers"

    text = f"{file_name} {sheet}"
    if "영업" in text or "사원" in text or "rep" in text:
        return "reps"
    if "판매" in text or "매장" in text or "store" in text:
        return "stores"
    if "대리점" in text or "dealer" in text:
        return "dealers"

    has_addr = bool(headers & STORE_ADDR_ALIASES)
    has_rep = bool(headers & REP_CODE_ALIASES)
    has_dealer_code = bool(headers & DEALER_CODE_ALIASES)
    has_dealer_name = bool(headers & DEALER_NAME_ALIASES)
    has_store_name = bool(headers & STORE_NAME_ALIASES)
    has_store_code = bool(headers & STORE_CODE_ALIASES)

    if has_store_code or has_addr or (has_store_name and not has_rep):
        return "stores"
    if has_rep:
        return "reps"
    if has_dealer_code or has_dealer_name:
        return "dealers"
    return None


def parse_uploads(files: list[tuple[str, bytes]]) -> dict[str, list[dict[str, str]]]:
    buckets: dict[str, list[dict[str, str]]] = {"dealers": [], "reps": [], "stores": [], "inventory": []}
    unknown_sheets: list[str] = []

    for filename, data in files:
        wb = load_workbook(BytesIO(data), data_only=True)
        try:
            for ws in wb.worksheets:
                rows = sheet_to_rows(ws)
                if not rows:
                    continue
                kind = guess_kind(filename, ws.title, set(rows[0].keys()))
                if kind == "inventory":
                    continue
                if not kind:
                    title_norm = ws.title.lower().replace(" ", "")
                    if "안내" in title_norm or "guide" in title_norm:
                        continue
                    unknown_sheets.append(f"{filename}:{ws.title}")
                    continue
                buckets[kind].extend(rows)
        finally:
            wb.close()

    buckets["_unknown"] = unknown_sheets  # type: ignore[assignment]
    return buckets


def _find_dealer(conn, code: str, name: str):
    if code:
        row = conn.execute("SELECT * FROM dealers WHERE dealer_code = ?", (code,)).fetchone()
        if row:
            return row
    if name:
        row = conn.execute("SELECT * FROM dealers WHERE name = ?", (name,)).fetchone()
        if row:
            return row
    return None


def upsert_masters(conn, buckets: dict[str, list[dict[str, str]]], now_iso: str, new_id) -> dict:
    summary = {
        "dealers": {"created": 0, "updated": 0, "skipped": 0, "errors": []},
        "reps": {"created": 0, "updated": 0, "skipped": 0, "errors": []},
        "stores": {"created": 0, "updated": 0, "skipped": 0, "duplicate_codes": 0, "errors": []},
        "unknown_sheets": buckets.get("_unknown", []),
    }

    for i, row in enumerate(buckets.get("dealers", []), start=2):
        code = pick(row, DEALER_CODE_ALIASES)
        name = pick(row, DEALER_NAME_ALIASES - {"소속대리점"}) or pick(row, {"대리점명", "대리점이름", "name", "dealername", "dealer_name"})
        if not name:
            name = pick(row, DEALER_NAME_ALIASES)
        if not code or not name:
            summary["dealers"]["skipped"] += 1
            summary["dealers"]["errors"].append(f"대리점 {i}행: 대리점ID/대리점명 필요")
            continue
        existing = conn.execute("SELECT * FROM dealers WHERE dealer_code = ?", (code,)).fetchone()
        if existing:
            conn.execute("UPDATE dealers SET name = ? WHERE id = ?", (name, existing["id"]))
            summary["dealers"]["updated"] += 1
        else:
            conn.execute(
                "INSERT INTO dealers (id, dealer_code, name, created_at) VALUES (?, ?, ?, ?)",
                (new_id(), code, name, now_iso),
            )
            summary["dealers"]["created"] += 1

    for i, row in enumerate(buckets.get("reps", []), start=2):
        code = pick(row, REP_CODE_ALIASES)
        name = pick(row, REP_NAME_ALIASES)
        dealer_code = pick(row, DEALER_CODE_ALIASES)
        dealer_name = pick(row, DEALER_NAME_ALIASES)
        if not code or not name:
            summary["reps"]["skipped"] += 1
            summary["reps"]["errors"].append(f"영업사원 {i}행: 고유ID/이름 필요")
            continue
        dealer = _find_dealer(conn, dealer_code, dealer_name)
        if not dealer:
            summary["reps"]["skipped"] += 1
            summary["reps"]["errors"].append(f"영업사원 {code}: 소속대리점을 찾을 수 없음 ({dealer_code or dealer_name or '빈값'})")
            continue
        existing = conn.execute("SELECT * FROM reps WHERE employee_code = ?", (code,)).fetchone()
        if existing:
            conn.execute(
                "UPDATE reps SET name = ?, dealer_id = ? WHERE id = ?",
                (name, dealer["id"], existing["id"]),
            )
            if not existing["password_hash"]:
                conn.execute(
                    "UPDATE reps SET password_hash = ? WHERE id = ?",
                    (generate_password_hash(code), existing["id"]),
                )
            summary["reps"]["updated"] += 1
        else:
            conn.execute(
                """
                INSERT INTO reps (id, dealer_id, name, employee_code, password_hash, device_id, created_at)
                VALUES (?, ?, ?, ?, ?, NULL, ?)
                """,
                (new_id(), dealer["id"], name, code, generate_password_hash(code), now_iso),
            )
            summary["reps"]["created"] += 1

    seen_codes: set[str] = set()
    for i, row in enumerate(buckets.get("stores", []), start=2):
        store_code = pick(row, STORE_CODE_ALIASES)
        address = pick(row, {"기본주소"}) or pick(row, STORE_ADDR_ALIASES)
        detail_address = pick(row, DETAIL_ADDR_ALIASES)
        name = pick(row, STORE_NAME_ALIASES) or address
        dealer_code = pick(row, DEALER_CODE_ALIASES)
        dealer_name = pick(row, DEALER_NAME_ALIASES)
        lat_raw = pick(row, LAT_ALIASES)
        lng_raw = pick(row, LNG_ALIASES)
        if not store_code:
            summary["stores"]["skipped"] += 1
            summary["stores"]["errors"].append(f"판매점 {i}행: 판매점코드 필요")
            continue
        if not address:
            summary["stores"]["skipped"] += 1
            summary["stores"]["errors"].append(f"판매점 {store_code}: 기본주소 필요")
            continue
        if store_code in seen_codes:
            summary["stores"]["duplicate_codes"] += 1
            continue
        seen_codes.add(store_code)

        dealer = _find_dealer(conn, dealer_code, dealer_name) if (dealer_code or dealer_name) else None
        try:
            lat = float(lat_raw) if lat_raw else 0.0
            lng = float(lng_raw) if lng_raw else 0.0
        except ValueError:
            summary["stores"]["skipped"] += 1
            summary["stores"]["errors"].append(f"판매점 {store_code}: 위도/경도 숫자 형식 오류")
            continue

        dealer_id = dealer["id"] if dealer else None
        existing = conn.execute("SELECT * FROM stores WHERE store_code = ?", (store_code,)).fetchone()
        if not existing:
            # 예전에 주소로 한 곳으로 합쳐 둔 행이 있으면, 그중 하나에 이 코드를 붙인다.
            existing = conn.execute(
                """
                SELECT * FROM stores
                WHERE address = ? AND (store_code IS NULL OR store_code = '')
                LIMIT 1
                """,
                (address,),
            ).fetchone()

        if existing:
            if lat == 0 and lng == 0:
                lat, lng = existing["lat"], existing["lng"]
            conn.execute(
                """
                UPDATE stores
                SET store_code = ?, name = ?, address = ?, detail_address = ?,
                    dealer_id = COALESCE(?, dealer_id), lat = ?, lng = ?
                WHERE id = ?
                """,
                (
                    store_code,
                    name,
                    address,
                    detail_address,
                    dealer_id,
                    lat,
                    lng,
                    existing["id"],
                ),
            )
            summary["stores"]["updated"] += 1
        else:
            conn.execute(
                """
                INSERT INTO stores (
                    id, dealer_id, store_code, name, address, detail_address, lat, lng, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    new_id(),
                    dealer_id,
                    store_code,
                    name,
                    address,
                    detail_address,
                    lat,
                    lng,
                    now_iso,
                ),
            )
            summary["stores"]["created"] += 1

    from geocode import copy_coords_for_same_address

    copy_coords_for_same_address(conn)

    return summary


def build_template_xlsx() -> bytes:
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    header_font = Font(color="FFFFFF", bold=True)
    example_fill = PatternFill("solid", fgColor="FEF3C7")

    sheets = [
        (
            "대리점",
            ["대리점ID", "대리점명"],
            [["DEAL001", "강남대리점"], ["DEAL002", "분당대리점"]],
            20,
        ),
        (
            "영업사원",
            ["고유ID", "이름", "소속대리점ID"],
            [["EMP001", "홍길동", "DEAL001"], ["EMP002", "김영업", "DEAL002"]],
            18,
        ),
        (
            "판매점",
            ["판매점코드", "판매점명", "기본주소", "상세주소", "소속대리점ID"],
            [
                ["PC9452", "테스트매장A", "서울특별시 중구 세종대로 110", "1층 101호", "DEAL001"],
                ["PE0840", "테스트매장B", "서울특별시 중구 세종대로 110", "1층 102호", "DEAL001"],
                ["P81790", "테스트매장C", "경기도 성남시 분당구 정자일로 95", "2층", "DEAL002"],
            ],
            36,
        ),
    ]

    for idx, (title, headers, examples, width) in enumerate(sheets):
        ws = wb.active if idx == 0 else wb.create_sheet(title)
        if idx == 0:
            ws.title = title
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(1, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col)].width = width
        for example in examples:
            ws.append(example)
            for col in range(1, len(headers) + 1):
                ws.cell(ws.max_row, col).fill = example_fill
        ws.freeze_panes = "A2"

    note = wb.create_sheet("작성안내")
    note["A1"] = "작성 안내"
    note["A1"].font = Font(bold=True, size=14)
    lines = [
        "1) 노란 행은 예시입니다. 지우고 실제 데이터를 넣으세요.",
        "2) 반드시 대리점 → 영업사원/판매점 순으로 연결됩니다. 소속대리점ID는 대리점 시트의 대리점ID와 같아야 합니다.",
        "3) 파일 3개로 나눠 올려도 되고, 이 엑셀 1개(시트 3개)로 올려도 됩니다.",
        "4) 판매점코드가 매장 고유키입니다. 주소가 같아도 코드가 다르면 다른 매장으로 등록합니다.",
        "5) GPS는 기본주소를 쓰고, 상세주소(층/호수)는 화면에 표시합니다.",
        "6) 보물찾기는 기본주소가 같으면 한 곳으로 봅니다. 판매점코드는 재고 구분용입니다.",
        "7) 판매점과 영업사원을 매칭하지 않습니다. 보물은 팀 전체가 볼 수 있습니다.",
        "8) 같은 대리점ID/고유ID/판매점코드는 다시 올리면 수정(업데이트)됩니다.",
        "9) .xlsx 형식만 지원합니다. 구형 .xls 는 엑셀에서 .xlsx 로 저장한 뒤 올려주세요.",
    ]
    for i, line in enumerate(lines, start=3):
        note[f"A{i}"] = line
    note.column_dimensions["A"].width = 110

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_stats_xlsx(conn) -> bytes:
    """전체 대리점/영업사원 포인트·방문 통계 엑셀."""
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    header_font = Font(color="FFFFFF", bold=True)

    dealer_rows = conn.execute(
        """
        SELECT
            d.dealer_code,
            d.name,
            (SELECT COUNT(*) FROM reps r WHERE r.dealer_id = d.id) AS rep_count,
            (
                SELECT COALESCE(SUM(pl.points), 0)
                FROM point_ledger pl
                JOIN reps r ON r.id = pl.rep_id
                WHERE r.dealer_id = d.id
            ) AS total_points,
            (
                SELECT COUNT(*)
                FROM visit_sessions vs
                JOIN reps r ON r.id = vs.rep_id
                WHERE r.dealer_id = d.id
            ) AS visit_count,
            (
                SELECT COUNT(*)
                FROM visit_sessions vs
                JOIN reps r ON r.id = vs.rep_id
                WHERE r.dealer_id = d.id AND vs.status = 'auto_approved'
            ) AS approved_visits
        FROM dealers d
        ORDER BY total_points DESC, d.name
        """
    ).fetchall()

    dealer_headers = [
        "대리점ID",
        "대리점명",
        "소속사원수",
        "총포인트",
        "방문시도",
        "인증성공",
    ]
    dealer_ws = wb.active
    dealer_ws.title = "대리점통계"
    dealer_ws.append(dealer_headers)
    for col in range(1, len(dealer_headers) + 1):
        cell = dealer_ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        dealer_ws.column_dimensions[get_column_letter(col)].width = 16
    for row in dealer_rows:
        dealer_ws.append(
            [
                row["dealer_code"],
                row["name"],
                int(row["rep_count"] or 0),
                int(row["total_points"] or 0),
                int(row["visit_count"] or 0),
                int(row["approved_visits"] or 0),
            ]
        )
    dealer_ws.freeze_panes = "A2"

    rep_rows = conn.execute(
        """
        SELECT
            r.employee_code,
            r.name,
            d.dealer_code,
            d.name AS dealer_name,
            COALESCE(SUM(pl.points), 0) AS total_points,
            (
                SELECT COUNT(*) FROM visit_sessions vs WHERE vs.rep_id = r.id
            ) AS visit_count,
            (
                SELECT COUNT(*) FROM visit_sessions vs
                WHERE vs.rep_id = r.id AND vs.status = 'auto_approved'
            ) AS approved_visits,
            MAX(pl.created_at) AS last_point_at
        FROM reps r
        LEFT JOIN dealers d ON d.id = r.dealer_id
        LEFT JOIN point_ledger pl ON pl.rep_id = r.id
        GROUP BY r.id
        ORDER BY total_points DESC, r.name
        """
    ).fetchall()

    rep_headers = [
        "고유ID",
        "이름",
        "소속대리점ID",
        "소속대리점명",
        "총포인트",
        "방문시도",
        "인증성공",
        "최근포인트일시",
    ]
    rep_ws = wb.create_sheet("영업사원통계")
    rep_ws.append(rep_headers)
    for col in range(1, len(rep_headers) + 1):
        cell = rep_ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        rep_ws.column_dimensions[get_column_letter(col)].width = 18
    for row in rep_rows:
        rep_ws.append(
            [
                row["employee_code"],
                row["name"],
                row["dealer_code"] or "",
                row["dealer_name"] or "소속 없음",
                int(row["total_points"] or 0),
                int(row["visit_count"] or 0),
                int(row["approved_visits"] or 0),
                row["last_point_at"] or "",
            ]
        )
    rep_ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
