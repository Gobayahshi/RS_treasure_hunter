"""재고현황 엑셀 파싱 및 판매점(P코드) 지도 집계."""

from __future__ import annotations

import csv
import re
from datetime import date, datetime
from io import BytesIO, StringIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from excel_import import cell_str, normalize_header
from confidence import haversine_distance_meters

HOLDER_CODE_ALIASES = {"보유처매장코드", "판매점코드", "매장코드"}
HOLDER_NAME_ALIASES = {"보유처", "판매점명", "매장명"}
PRODUCT_SHORT_ALIASES = {"대표상품명", "대표모델명"}
MODEL_ALIASES = {"모델명"}
PRICE_ALIASES = {"실구매가", "구매가격"}
INBOUND_ALIASES = {"입고일자"}
MOVED_ALIASES = {"재고이동출고일자"}
HOLD_DAYS_ALIASES = {"보유기간"}
SERIAL_ALIASES = {"일련번호"}
DEALER_ORG_ALIASES = {"레벨0조직명", "대리점명"}
DEALER_CODE_IN_FILENAME = re.compile(r"(D\d{5})", re.I)
AGED_DAYS = 30
HOLD_WARN_DAYS = 15
DEFAULT_MAP_MODELS = ("SM-F971", "SM-A175N", "SM-S931N")
_OVERVIEW_CACHE: dict[tuple, dict] = {}


def clear_inventory_cache() -> None:
    _OVERVIEW_CACHE.clear()


def classify_holder(store_code: str) -> str:
    code = (store_code or "").strip().upper()
    if code.startswith("P"):
        return "partner"
    if code.startswith("D") and len(code) > 6:
        return "retail"
    if code.startswith("D"):
        return "hq"
    return "other"


# 긴 접두어를 먼저 본다. 주소.xlsx 기본주소(서울특별시 …)와 시도명(서울) 둘 다 받는다.
REGION_PREFIXES = (
    ("서울", ("서울특별시", "서울시", "서울")),
    ("인천", ("인천광역시", "인천시", "인천")),
    ("경기", ("경기도", "경기")),
    ("부산", ("부산광역시", "부산시", "부산")),
    ("대구", ("대구광역시", "대구시", "대구")),
    ("광주", ("광주광역시", "광주시", "광주")),
    ("대전", ("대전광역시", "대전시", "대전")),
    ("울산", ("울산광역시", "울산시", "울산")),
    ("세종", ("세종특별자치시", "세종시", "세종")),
    ("강원", ("강원특별자치도", "강원도", "강원")),
    ("충북", ("충청북도", "충북")),
    ("충남", ("충청남도", "충남")),
    ("전북", ("전북특별자치도", "전라북도", "전북")),
    ("전남", ("전라남도", "전남")),
    ("경북", ("경상북도", "경북")),
    ("경남", ("경상남도", "경남")),
    ("제주", ("제주특별자치도", "제주도", "제주")),
)


def address_region(address: str) -> str:
    text = (address or "").strip().replace(" ", "")
    if not text:
        return "기타"
    for region, prefixes in REGION_PREFIXES:
        for prefix in prefixes:
            if text.startswith(prefix.replace(" ", "")):
                return region
    return "기타"


def _pick(row: dict[str, str], aliases: set[str]) -> str:
    for key, value in row.items():
        if key in aliases and value:
            return value
    return ""


def _find_inventory_header(ws: Worksheet) -> tuple[int, dict[str, int]]:
    for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        mapping = {}
        for idx, value in enumerate(row, start=1):
            key = normalize_header(value)
            if key:
                mapping[key] = idx
        if mapping.keys() & HOLDER_CODE_ALIASES and (
            mapping.keys() & PRODUCT_SHORT_ALIASES or mapping.keys() & MODEL_ALIASES
        ):
            return row_num, mapping
    return 0, {}


def _parse_as_of_value(val) -> str:
    if val is None or val == "":
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y%m%d")
    if isinstance(val, date) and not isinstance(val, datetime):
        return val.strftime("%Y%m%d")
    if isinstance(val, bool):
        return ""
    if isinstance(val, (int, float)):
        number = int(val)
        if 20200101 <= number <= 20991231:
            return str(number)
        return ""
    text = cell_str(val)
    match = re.search(r"일자[:\s]*([0-9]{4})[.\-/]?([0-9]{2})[.\-/]?([0-9]{2})", text)
    if match:
        return "".join(match.groups())
    digits = re.sub(r"\D", "", text)
    if len(digits) == 8 and digits.startswith("20"):
        return digits
    return ""


def _as_of_date(ws: Worksheet) -> str:
    """재고 파일 A2, 또는 헤더보다 위 칸의 기준일을 YYYYMMDD로 읽는다."""
    try:
        parsed = _parse_as_of_value(ws.cell(2, 1).value)
        if parsed:
            return parsed
    except Exception:
        pass
    header_row, _ = _find_inventory_header(ws)
    top = header_row - 1 if header_row and header_row > 1 else 0
    if top < 1:
        return ""
    for row in ws.iter_rows(min_row=1, max_row=top, max_col=20, values_only=True):
        for val in row:
            parsed = _parse_as_of_value(val)
            if parsed:
                return parsed
    return ""


def is_inventory_workbook(data: bytes) -> bool:
    wb = load_workbook(BytesIO(data), data_only=True)
    try:
        for ws in wb.worksheets:
            _, headers = _find_inventory_header(ws)
            if headers:
                return True
        return False
    finally:
        wb.close()


def _item_from_raw(raw: dict[str, str]) -> dict[str, str] | None:
    store_code = _pick(raw, HOLDER_CODE_ALIASES).strip().upper()
    if not store_code:
        return None
    hold_raw = _pick(raw, HOLD_DAYS_ALIASES)
    try:
        hold_days = str(int(float(hold_raw))) if hold_raw else ""
    except ValueError:
        hold_days = hold_raw
    return {
        "store_code": store_code,
        "holder_name": _pick(raw, HOLDER_NAME_ALIASES),
        "holder_type": classify_holder(store_code),
        "product_short": _pick(raw, PRODUCT_SHORT_ALIASES),
        "model_name": _pick(raw, MODEL_ALIASES),
        "purchase_price": _pick(raw, PRICE_ALIASES),
        "inbound_date": _pick(raw, INBOUND_ALIASES)[:10],
        "moved_date": _pick(raw, MOVED_ALIASES)[:10],
        "hold_days": hold_days,
        "serial": _pick(raw, SERIAL_ALIASES),
        "dealer_org": _pick(raw, DEALER_ORG_ALIASES),
    }


def _decode_csv_text(data: bytes) -> str:
    for enc in ("utf-8-sig", "cp949", "utf-8"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def parse_inventory_csv(filename: str, data: bytes) -> dict[str, Any]:
    lines = list(csv.reader(StringIO(_decode_csv_text(data))))
    header_idx = -1
    headers: dict[str, int] = {}
    as_of = ""
    for i, row in enumerate(lines[:25]):
        mapping = {}
        for idx, value in enumerate(row, start=1):
            key = normalize_header(value)
            if key:
                mapping[key] = idx
        if mapping.keys() & HOLDER_CODE_ALIASES and (
            mapping.keys() & PRODUCT_SHORT_ALIASES or mapping.keys() & MODEL_ALIASES
        ):
            header_idx = i
            headers = mapping
            break
        for val in row:
            as_of = as_of or _parse_as_of_value(val)
    if header_idx < 0:
        return {"filename": filename, "as_of_date": as_of, "rows": []}
    rows: list[dict[str, str]] = []
    for row in lines[header_idx + 1 :]:
        raw = {h: cell_str(row[col - 1] if col - 1 < len(row) else "") for h, col in headers.items()}
        if not any(raw.values()):
            continue
        item = _item_from_raw(raw)
        if item:
            rows.append(item)
    return {"filename": filename, "as_of_date": as_of, "rows": rows}


def is_inventory_csv(data: bytes) -> bool:
    parsed = parse_inventory_csv("check.csv", data)
    return bool(parsed.get("rows"))


def parse_inventory_file(filename: str, data: bytes) -> dict[str, Any]:
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return parse_inventory_csv(filename, data)
    return parse_inventory_xlsx(filename, data)


def parse_inventory_xlsx(filename: str, data: bytes) -> dict[str, Any]:
    wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
    try:
        rows: list[dict[str, str]] = []
        as_of = ""
        for ws in wb.worksheets:
            headers: dict[str, int] = {}
            scanned = 0
            for row in ws.iter_rows(values_only=True):
                row = row or ()
                if not headers:
                    scanned += 1
                    if not as_of:
                        for val in row[:20]:
                            as_of = as_of or _parse_as_of_value(val)
                    mapping = {}
                    for idx, value in enumerate(row, start=1):
                        key = normalize_header(value)
                        if key:
                            mapping[key] = idx
                    if mapping.keys() & HOLDER_CODE_ALIASES and (
                        mapping.keys() & PRODUCT_SHORT_ALIASES or mapping.keys() & MODEL_ALIASES
                    ):
                        headers = mapping
                    elif scanned >= 20:
                        break
                    continue
                raw = {
                    h: cell_str(row[col - 1] if col - 1 < len(row) else None)
                    for h, col in headers.items()
                }
                if not any(raw.values()):
                    continue
                item = _item_from_raw(raw)
                if item:
                    rows.append(item)
        return {"filename": filename, "as_of_date": as_of, "rows": rows}
    finally:
        wb.close()


def _hold_days_int(value) -> int | None:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def infer_dealer(conn, filename: str, rows: list[dict[str, str]] | None = None):
    """파일명 D코드 또는 엑셀 레벨0조직명으로 대리점을 찾는다."""
    code = ""
    match = DEALER_CODE_IN_FILENAME.search(filename or "")
    if match:
        code = match.group(1).upper()
        row = conn.execute("SELECT * FROM dealers WHERE dealer_code = ?", (code,)).fetchone()
        if row:
            return dict(row)

    name_hints: list[str] = []
    if match:
        rest = (filename or "")[match.end() :]
        rest = re.split(r"[_\s.]", rest.lstrip(" _-"))[0].strip()
        if rest and rest.lower() not in {"실시간재고현황", "재고현황", "xlsx"}:
            name_hints.append(rest)
    counts: dict[str, int] = {}
    for item in rows or []:
        org = (item.get("dealer_org") or "").strip()
        if org:
            counts[org] = counts.get(org, 0) + 1
    if counts:
        name_hints.insert(0, max(counts, key=counts.get))

    dealers = [dict(r) for r in conn.execute("SELECT * FROM dealers").fetchall()]
    for hint in name_hints:
        compact = hint.replace(" ", "")
        for dealer in dealers:
            name = (dealer.get("name") or "").replace(" ", "")
            if not name:
                continue
            if name in compact or compact in name or compact.startswith(name):
                return dealer

    if code:
        name = name_hints[0] if name_hints else code
        existing = conn.execute("SELECT * FROM dealers WHERE dealer_code = ?", (code,)).fetchone()
        if existing:
            return dict(existing)
        from datetime import datetime

        dealer_id = __import__("uuid").uuid4().hex
        conn.execute(
            "INSERT INTO dealers (id, dealer_code, name, created_at) VALUES (?, ?, ?, ?)",
            (dealer_id, code, name, datetime.utcnow().isoformat()),
        )
        return dict(conn.execute("SELECT * FROM dealers WHERE id = ?", (dealer_id,)).fetchone())
    return None


def replace_inventory(conn, parsed: dict[str, Any], now_iso: str, new_id, dealer=None) -> dict:
    """해당 대리점의 이전 재고는 지우고 이번 파일만 남긴다."""
    clear_inventory_cache()
    rows = parsed.get("rows") or []
    if dealer is None:
        dealer = infer_dealer(conn, parsed.get("filename") or "", rows)
    elif not isinstance(dealer, dict):
        dealer = dict(dealer)
    if not dealer:
        raise ValueError("어느 대리점 재고인지 모릅니다. 파일명에 대리점코드(예: D14746)를 넣어주세요.")

    old = conn.execute(
        "SELECT id FROM inventory_uploads WHERE dealer_id = ?", (dealer["id"],)
    ).fetchall()
    old_ids = [r["id"] for r in old]
    if old_ids:
        placeholders = ",".join("?" * len(old_ids))
        conn.execute(f"DELETE FROM inventory_items WHERE upload_id IN ({placeholders})", old_ids)
        conn.execute(f"DELETE FROM inventory_uploads WHERE id IN ({placeholders})", old_ids)
    conn.execute("DELETE FROM inventory_items WHERE dealer_id = ?", (dealer["id"],))

    upload_id = new_id()
    conn.execute(
        """
        INSERT INTO inventory_uploads (
            id, filename, as_of_date, row_count, created_at, dealer_id, dealer_code, dealer_name
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            upload_id,
            parsed.get("filename") or "",
            parsed.get("as_of_date") or "",
            len(rows),
            now_iso,
            dealer["id"],
            dealer.get("dealer_code") or "",
            dealer.get("name") or "",
        ),
    )
    conn.executemany(
        """
        INSERT INTO inventory_items (
            id, upload_id, store_code, holder_name, holder_type,
            product_short, model_name, purchase_price, inbound_date,
            moved_date, hold_days, serial, dealer_id, dealer_code, dealer_name
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [
            (
                new_id(),
                upload_id,
                row["store_code"],
                row["holder_name"],
                row["holder_type"],
                row["product_short"],
                row["model_name"],
                row["purchase_price"],
                row["inbound_date"],
                row["moved_date"],
                _hold_days_int(row["hold_days"]),
                row["serial"],
                dealer["id"],
                dealer.get("dealer_code") or "",
                dealer.get("name") or "",
            )
            for row in rows
        ],
    )
    by_type: dict[str, int] = {}
    for row in rows:
        by_type[row["holder_type"]] = by_type.get(row["holder_type"], 0) + 1
    return {
        "upload_id": upload_id,
        "filename": parsed.get("filename") or "",
        "as_of_date": parsed.get("as_of_date") or "",
        "row_count": len(rows),
        "by_holder_type": by_type,
        "dealer_id": dealer["id"],
        "dealer_code": dealer.get("dealer_code") or "",
        "dealer_name": dealer.get("name") or "",
    }


def _canon_model_token(token: str) -> str:
    text = (token or "").strip().upper().replace(" ", "")
    if not text:
        return ""
    if text.startswith("SM") and not text.startswith("SM-"):
        return "SM-" + text[2:]
    return text


def _values_list(value) -> list[str]:
    if value is None:
        return []
    if isinstance(value, (list, tuple, set)):
        raw = list(value)
    else:
        raw = re.split(r"[,|]+", str(value))
    out: list[str] = []
    seen: set[str] = set()
    for item in raw:
        text = str(item).strip()
        if not text:
            continue
        key = text.upper()
        if key in seen:
            continue
        seen.add(key)
        out.append(text)
    return out


def parse_models(model_prefix: str | None) -> list[str]:
    """빈 값·ALL 이면 전체 기종. 예전처럼 3기종으로 기본 고정하지 않는다."""
    raw = (model_prefix or "").strip()
    compact = raw.upper().replace(" ", "")
    if compact in {"", "ALL", "*", "ALLMODELS"}:
        return []
    parts = []
    for token in re.split(r"[,|]+", raw):
        token = token.strip()
        if not token:
            continue
        if token.upper().replace(" ", "") in {"ALL", "*"}:
            return []
        canon = _canon_model_token(token)
        if canon and canon not in parts:
            parts.append(canon)
    return parts


def _has_coords(lat, lng) -> bool:
    try:
        return float(lat or 0) != 0 or float(lng or 0) != 0
    except (TypeError, ValueError):
        return False


def _empty_map(
    model: str,
    include_retail: bool,
    region: str,
    keyword: str,
    dealer_id: str = "",
) -> dict:
    return {
        "model": model,
        "as_of_date": "",
        "filename": "",
        "include_retail": include_retail,
        "region": region,
        "keyword": keyword,
        "dealer_id": dealer_id,
        "product_short": "",
        "model_name": "",
        "pin_color": "",
        "color_mode": "hold",
        "total_qty": 0,
        "mapped_qty": 0,
        "unmapped_qty": 0,
        "store_count": 0,
        "points": [],
        "unmapped": [],
        "regions": [],
        "dealer_totals": [],
        "uploads": [],
        "nearest": None,
        "aged_days": AGED_DAYS,
        "shared_store_count": 0,
        "bbox": None,
        "aged_only": False,
        "area_model_totals": [],
    }


def _latest_upload_ids(conn, dealer_id: str | None = None) -> list[str]:
    if dealer_id:
        row = conn.execute(
            """
            SELECT id FROM inventory_uploads
            WHERE dealer_id = ?
            ORDER BY created_at DESC
            LIMIT 1
            """,
            (dealer_id,),
        ).fetchone()
        return [row["id"]] if row else []
    rows = conn.execute(
        """
        SELECT u.id
        FROM inventory_uploads u
        JOIN (
            SELECT COALESCE(dealer_id, id) AS grp, MAX(created_at) AS created_at
            FROM inventory_uploads
            GROUP BY COALESCE(dealer_id, id)
        ) latest
          ON COALESCE(u.dealer_id, u.id) = latest.grp
         AND u.created_at = latest.created_at
        """
    ).fetchall()
    return [r["id"] for r in rows]


def _store_display_name(row, code: str) -> str:
    store_name = (row["store_name"] or "").strip()
    holder_name = (row["holder_name"] or "").strip()
    address = (row["address"] or "").strip()
    if store_name and store_name != address:
        return store_name
    if holder_name and holder_name != address:
        return holder_name
    return store_name or holder_name or code


def _merge_store_rows(rows) -> list[dict]:
    grouped: dict[str, dict] = {}
    for row in rows:
        raw_code = row["store_code"]
        list_code = ""
        if "list_store_code" in row.keys():
            list_code = (row["list_store_code"] or "").strip()
        code = list_code or raw_code
        qty = int(row["qty"] or 0)
        item = grouped.get(raw_code)
        if not item:
            address = row["address"] or ""
            item = {
                "store_code": code,
                "name": _store_display_name(row, code),
                "holder_name": row["holder_name"] or "",
                "address": address,
                "detail_address": row["detail_address"] or "",
                "region": address_region(address),
                "lat": row["lat"],
                "lng": row["lng"],
                "qty": 0,
                "avg_hold_days": None,
                "max_hold_days": None,
                "aged_qty": 0,
                "_dealers": {},
                "_models": {},
            }
            grouped[raw_code] = item
        hold_sum = (item.get("_hold_sum") or 0) + (float(row["avg_hold_days"] or 0) * qty)
        item["_hold_sum"] = hold_sum
        item["qty"] += qty
        item["aged_qty"] += int(row["aged_qty"] or 0)
        max_hold = row["max_hold_days"]
        if max_hold is not None:
            prev = item["max_hold_days"]
            item["max_hold_days"] = int(max_hold) if prev is None else max(int(prev), int(max_hold))
        dealer_key = row["dealer_id"] or row["dealer_code"] or row["dealer_name"] or "unknown"
        dealer = item["_dealers"].get(dealer_key)
        if not dealer:
            dealer = {
                "dealer_id": row["dealer_id"] or "",
                "dealer_code": row["dealer_code"] or "",
                "dealer_name": row["dealer_name"] or "미지정",
                "qty": 0,
                "avg_hold_days": None,
                "max_hold_days": None,
                "aged_qty": 0,
            }
            item["_dealers"][dealer_key] = dealer
        dealer["qty"] += qty
        dealer["aged_qty"] += int(row["aged_qty"] or 0)
        if row["max_hold_days"] is not None:
            prev = dealer["max_hold_days"]
            dealer["max_hold_days"] = (
                int(row["max_hold_days"]) if prev is None else max(int(prev), int(row["max_hold_days"]))
            )
        model_key = (row["model_key"] if "model_key" in row.keys() else "") or ""
        if model_key:
            item["_models"][model_key] = item["_models"].get(model_key, 0) + qty
        if row["holder_name"]:
            item["holder_name"] = row["holder_name"]
        display = _store_display_name(row, code)
        if display:
            item["name"] = display
        if list_code:
            item["store_code"] = list_code
    out = []
    for item in grouped.values():
        hold_sum = item.pop("_hold_sum", 0)
        dealers_map = item.pop("_dealers")
        models_map = item.pop("_models")
        if item["qty"]:
            item["avg_hold_days"] = round(hold_sum / item["qty"], 1)
        item["dealers"] = sorted(dealers_map.values(), key=lambda d: (-d["qty"], d["dealer_name"]))
        item["models"] = [
            {"model": name, "qty": qty}
            for name, qty in sorted(models_map.items(), key=lambda kv: (-kv[1], kv[0]))
        ]
        item["shared"] = len(item["dealers"]) > 1
        out.append(item)
    out.sort(key=lambda p: (-p["qty"], p["store_code"]))
    return out


def normalize_bbox(bbox) -> tuple[float, float, float, float] | None:
    if not bbox:
        return None
    if isinstance(bbox, dict):
        try:
            south = float(bbox.get("south"))
            west = float(bbox.get("west"))
            north = float(bbox.get("north"))
            east = float(bbox.get("east"))
        except (TypeError, ValueError):
            return None
    elif isinstance(bbox, (list, tuple)) and len(bbox) == 4:
        try:
            south, west, north, east = (float(x) for x in bbox)
        except (TypeError, ValueError):
            return None
    else:
        return None
    if south > north:
        south, north = north, south
    if west > east:
        west, east = east, west
    if south == north or west == east:
        return None
    return south, west, north, east


def _in_bbox(lat, lng, bbox: tuple[float, float, float, float]) -> bool:
    south, west, north, east = bbox
    try:
        y = float(lat)
        x = float(lng)
    except (TypeError, ValueError):
        return False
    return south <= y <= north and west <= x <= east


def inventory_map_points(
    conn,
    model_prefix: str,
    include_retail: bool = False,
    region: str | None = None,
    lat: float | None = None,
    lng: float | None = None,
    keyword: str | None = None,
    dealer_id: str | None = None,
    bbox=None,
    aged_only: bool = False,
    radius_km: float | None = None,
    product_short: str | None = None,
    model_name: str | None = None,
    pin_color: str | None = None,
) -> dict:
    wanted_shorts = _values_list(product_short)
    wanted_model_names = _values_list(model_name)
    models = [] if (wanted_shorts or wanted_model_names) else parse_models(model_prefix)
    if wanted_model_names:
        model = ",".join(wanted_model_names)
    elif wanted_shorts:
        model = ",".join(wanted_shorts)
    else:
        model = ",".join(models) if models else "all"
    holders = ("partner", "retail") if include_retail else ("partner",)
    placeholders = ",".join("?" * len(holders))
    wanted_region = (region or "").strip()
    wanted_keyword = (keyword or "").strip()
    wanted_dealer = (dealer_id or "").strip()
    wanted_bbox = normalize_bbox(bbox)
    try:
        radius_m = float(radius_km) * 1000 if radius_km not in (None, "") else None
    except (TypeError, ValueError):
        radius_m = None
    if radius_m is not None and radius_m <= 0:
        radius_m = None

    upload_ids = _latest_upload_ids(conn, wanted_dealer or None)
    if not upload_ids:
        return _empty_map(model, include_retail, wanted_region, wanted_keyword, wanted_dealer)

    up_ph = ",".join("?" * len(upload_ids))
    model_wheres = []
    model_params: list = []
    case_sql = []
    case_params: list = []
    if wanted_model_names:
        ph = ",".join("?" * len(wanted_model_names))
        model_filter_sql = f"AND UPPER(TRIM(COALESCE(i.model_name, ''))) IN ({ph})"
        model_params.extend([name.upper() for name in wanted_model_names])
        model_key_expr = "UPPER(TRIM(COALESCE(i.model_name, i.product_short, '')))"
    elif wanted_shorts:
        ph = ",".join("?" * len(wanted_shorts))
        model_filter_sql = f"AND UPPER(TRIM(COALESCE(i.product_short, ''))) IN ({ph})"
        model_params.extend([name.upper() for name in wanted_shorts])
        model_key_expr = "UPPER(TRIM(COALESCE(NULLIF(i.model_name, ''), i.product_short, '')))"
    elif models:
        for name in models:
            like = f"{name}%"
            model_wheres.append(
                "(UPPER(COALESCE(i.product_short, '')) LIKE ? OR UPPER(COALESCE(i.model_name, '')) LIKE ?)"
            )
            model_params.extend([like, like])
            case_sql.append(
                "WHEN UPPER(COALESCE(i.product_short, '')) LIKE ? OR UPPER(COALESCE(i.model_name, '')) LIKE ? THEN ?"
            )
            case_params.extend([like, like, name])
        model_key_expr = "CASE " + " ".join(case_sql) + " ELSE UPPER(COALESCE(i.product_short, i.model_name, '')) END"
        model_filter_sql = "AND (" + " OR ".join(model_wheres) + ")"
    else:
        model_key_expr = "UPPER(COALESCE(NULLIF(TRIM(i.product_short), ''), TRIM(i.model_name), ''))"
        model_filter_sql = ""
    rows = conn.execute(
        f"""
        SELECT
            i.store_code,
            i.dealer_id,
            MAX(i.dealer_code) AS dealer_code,
            MAX(i.dealer_name) AS dealer_name,
            MAX(i.holder_name) AS holder_name,
            {model_key_expr} AS model_key,
            COUNT(*) AS qty,
            AVG(i.hold_days) AS avg_hold_days,
            MAX(i.hold_days) AS max_hold_days,
            SUM(CASE WHEN i.hold_days >= ? THEN 1 ELSE 0 END) AS aged_qty,
            MAX(s.store_code) AS list_store_code,
            MAX(s.name) AS store_name,
            MAX(s.address) AS address,
            MAX(s.detail_address) AS detail_address,
            MAX(s.lat) AS lat,
            MAX(s.lng) AS lng
        FROM inventory_items i
        LEFT JOIN stores s ON UPPER(TRIM(COALESCE(s.store_code, ''))) = UPPER(TRIM(COALESCE(i.store_code, '')))
        WHERE i.upload_id IN ({up_ph})
          AND i.holder_type IN ({placeholders})
          {model_filter_sql}
        GROUP BY i.store_code, i.dealer_id, model_key
        { "HAVING aged_qty > 0" if aged_only else "" }
        """,
        (*case_params, AGED_DAYS, *upload_ids, *holders, *model_params),
    ).fetchall()

    merged = _merge_store_rows(rows)
    points = []
    unmapped = []
    for item in merged:
        if _has_coords(item["lat"], item["lng"]):
            points.append(item)
        else:
            unmapped.append(item)

    region_counts: dict[str, dict[str, int]] = {}
    for item in points:
        bucket = region_counts.setdefault(item["region"], {"qty": 0, "stores": 0})
        bucket["qty"] += item["qty"]
        bucket["stores"] += 1
    regions = [
        {"region": name, "qty": stats["qty"], "stores": stats["stores"]}
        for name, stats in sorted(region_counts.items(), key=lambda kv: (-kv[1]["qty"], kv[0]))
    ]

    if wanted_region:
        points = [p for p in points if p["region"] == wanted_region]
        unmapped = [u for u in unmapped if u["region"] == wanted_region]
    if wanted_keyword:
        key = wanted_keyword.replace(" ", "").lower()

        def _hit(item):
            dealer_blob = "".join(d.get("dealer_name") or "" for d in item.get("dealers") or [])
            blob = f"{item.get('name','')}{item.get('holder_name','')}{item.get('address','')}{item.get('store_code','')}{dealer_blob}"
            return key in blob.replace(" ", "").lower()

        points = [p for p in points if _hit(p)]
        unmapped = [u for u in unmapped if _hit(u)]

    if wanted_bbox:
        points = [p for p in points if _in_bbox(p.get("lat"), p.get("lng"), wanted_bbox)]
        unmapped = []
    if aged_only:
        points = [p for p in points if (p.get("aged_qty") or 0) > 0]
        unmapped = [u for u in unmapped if (u.get("aged_qty") or 0) > 0]

    nearest = None
    if lat is not None and lng is not None:
        for item in points:
            item["distance_meters"] = round(
                haversine_distance_meters(lat, lng, float(item["lat"]), float(item["lng"]))
            )
        points.sort(key=lambda p: (p.get("distance_meters") if p.get("distance_meters") is not None else 10**12, -p["qty"]))
        if radius_m is not None:
            nearby = [p for p in points if (p.get("distance_meters") or 10**12) <= radius_m]
            if nearby:
                points = nearby
        if points:
            nearest = dict(points[0])

    mapped_qty = sum(p["qty"] for p in points)
    unmapped_qty = sum(u["qty"] for u in unmapped)
    dealer_totals: dict[str, dict] = {}
    for item in points + unmapped:
        for dealer in item.get("dealers") or []:
            key = dealer.get("dealer_id") or dealer.get("dealer_code") or dealer.get("dealer_name")
            bucket = dealer_totals.setdefault(
                key,
                {
                    "dealer_id": dealer.get("dealer_id") or "",
                    "dealer_code": dealer.get("dealer_code") or "",
                    "dealer_name": dealer.get("dealer_name") or "",
                    "qty": 0,
                    "stores": 0,
                },
            )
            bucket["qty"] += dealer["qty"]
            bucket["stores"] += 1
    uploads = [
        dict(r)
        for r in conn.execute(
            f"SELECT filename, as_of_date, row_count, dealer_code, dealer_name FROM inventory_uploads WHERE id IN ({up_ph})",
            upload_ids,
        ).fetchall()
    ]
    as_of_dates = sorted({u.get("as_of_date") or "" for u in uploads if u.get("as_of_date")})
    shared_store_count = sum(1 for p in points if p.get("shared"))
    model_totals: dict[str, dict] = {name: {"model": name, "qty": 0, "stores": 0} for name in models}
    for item in points + unmapped:
        seen = set()
        for m in item.get("models") or []:
            bucket = model_totals.setdefault(m["model"], {"model": m["model"], "qty": 0, "stores": 0})
            bucket["qty"] += m["qty"]
            if m["model"] not in seen:
                bucket["stores"] += 1
                seen.add(m["model"])

    return {
        "model": model,
        "models": models,
        "product_short": ",".join(wanted_shorts),
        "model_name": ",".join(wanted_model_names),
        "pin_color": (pin_color or "").strip(),
        "color_mode": "custom" if (pin_color or "").strip() else "hold",
        "model_totals": [model_totals[name] for name in models if name in model_totals],
        "as_of_date": ",".join(as_of_dates),
        "filename": ", ".join(u.get("filename") or "" for u in uploads),
        "include_retail": include_retail,
        "region": wanted_region,
        "keyword": wanted_keyword,
        "dealer_id": wanted_dealer,
        "total_qty": mapped_qty + unmapped_qty,
        "store_count": len(points) + len(unmapped),
        "mapped_qty": mapped_qty,
        "unmapped_qty": unmapped_qty,
        "points": points,
        "unmapped": unmapped,
        "regions": regions,
        "dealer_totals": sorted(dealer_totals.values(), key=lambda d: (-d["qty"], d["dealer_name"])),
        "uploads": uploads,
        "nearest": nearest,
        "aged_days": AGED_DAYS,
        "shared_store_count": shared_store_count,
        "bbox": (
            {
                "south": wanted_bbox[0],
                "west": wanted_bbox[1],
                "north": wanted_bbox[2],
                "east": wanted_bbox[3],
            }
            if wanted_bbox
            else None
        ),
        "aged_only": bool(aged_only),
        "radius_km": (radius_m / 1000) if radius_m else None,
        "area_model_totals": sorted(
            model_totals.values(), key=lambda m: (-m["qty"], m["model"])
        ),
    }


def _parse_money(value) -> float | None:
    text = re.sub(r"[^\d.\-]", "", str(value or "").strip())
    if not text or text in {".", "-", "-."}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def inventory_store_price_sum(conn, store_code: str, dealer_id: str | None = None) -> dict:
    """한 판매점 재고의 실구매가 합계."""
    upload_ids, _uploads = _partner_upload_filter(conn, dealer_id)
    code = re.sub(r"\s+", "", (store_code or "")).upper()
    empty = {
        "store_code": code,
        "name": "",
        "address": "",
        "qty": 0,
        "total_price": 0,
        "missing_price": 0,
        "by_model": [],
    }
    if not upload_ids or not code:
        return empty
    rows = conn.execute(
        f"""
        SELECT
            i.purchase_price,
            COALESCE(NULLIF(TRIM(i.model_name), ''), NULLIF(TRIM(i.product_short), ''), '미상') AS model_name,
            s.name AS store_name,
            i.holder_name,
            s.address
        FROM inventory_items i
        LEFT JOIN stores s
          ON UPPER(TRIM(COALESCE(s.store_code, ''))) = UPPER(TRIM(COALESCE(i.store_code, '')))
        WHERE i.upload_id IN ({",".join("?" * len(upload_ids))})
          AND i.holder_type = 'partner'
          AND UPPER(REPLACE(TRIM(COALESCE(i.store_code, '')), ' ', '')) = ?
        """,
        (*upload_ids, code),
    ).fetchall()
    by_model: dict[str, dict] = {}
    total = 0.0
    missing = 0
    name = ""
    address = ""
    for row in rows:
        name = name or (row["store_name"] or row["holder_name"] or "")
        address = address or (row["address"] or "")
        model = row["model_name"] or "미상"
        item = by_model.setdefault(model, {"model": model, "qty": 0, "amount": 0})
        item["qty"] += 1
        amount = _parse_money(row["purchase_price"])
        if amount is None:
            missing += 1
        else:
            item["amount"] += amount
            total += amount
    if name and address and name.replace(" ", "") == address.replace(" ", ""):
        name = ""
    return {
        "store_code": code,
        "name": name or code,
        "address": address,
        "qty": len(rows),
        "total_price": total,
        "missing_price": missing,
        "by_model": sorted(by_model.values(), key=lambda m: m["model"]),
    }
    upload_ids = _latest_upload_ids(conn, dealer_id or None)
    if not upload_ids:
        return [], []
    uploads = [
        dict(r)
        for r in conn.execute(
            f"SELECT dealer_id, filename, as_of_date, row_count, created_at, dealer_code, dealer_name FROM inventory_uploads WHERE id IN ({','.join('?' * len(upload_ids))})",
            upload_ids,
        ).fetchall()
    ]
    return upload_ids, uploads


def _region_where(region: str) -> tuple[str, list]:
    wanted = (region or "").strip()
    if not wanted:
        return "", []
    prefixes = []
    for name, opts in REGION_PREFIXES:
        if name == wanted:
            prefixes = list(opts)
            break
    if not prefixes:
        return "", []
    clauses = []
    params: list = []
    for prefix in prefixes:
        clauses.append("REPLACE(COALESCE(s.address, ''), ' ', '') LIKE ?")
        params.append(prefix.replace(" ", "") + "%")
    return "(" + " OR ".join(clauses) + ")", params


def _scope_filters(region: str | None, keyword: str | None, bbox) -> tuple[str, list]:
    sql = ""
    params: list = []
    region_sql, region_params = _region_where(region or "")
    if region_sql:
        sql += f" AND {region_sql}"
        params.extend(region_params)
    wanted_bbox = normalize_bbox(bbox)
    if wanted_bbox:
        south, west, north, east = wanted_bbox
        sql += " AND s.lat IS NOT NULL AND s.lng IS NOT NULL AND s.lat BETWEEN ? AND ? AND s.lng BETWEEN ? AND ?"
        params.extend([south, north, west, east])
    key = (keyword or "").strip()
    if key:
        like = f"%{key}%"
        sql += """ AND (
            COALESCE(s.name, '') LIKE ?
            OR COALESCE(s.address, '') LIKE ?
            OR COALESCE(i.holder_name, '') LIKE ?
            OR COALESCE(i.store_code, '') LIKE ?
            OR COALESCE(i.dealer_name, '') LIKE ?
        )"""
        params.extend([like, like, like, like, like])
    return sql, params


def inventory_model_catalog(conn, dealer_id: str | None = None) -> dict:
    """드롭다운용 대표상품명 → 모델명 목록. 판매점(P코드) 재고만."""
    upload_ids, uploads = _partner_upload_filter(conn, dealer_id)
    as_of_dates = sorted({u.get("as_of_date") or "" for u in uploads if u.get("as_of_date")})
    empty = {
        "as_of_date": ",".join(as_of_dates),
        "uploads": uploads,
        "products": [],
    }
    if not upload_ids:
        return empty
    up_ph = ",".join("?" * len(upload_ids))
    rows = conn.execute(
        f"""
        SELECT
            COALESCE(NULLIF(TRIM(i.product_short), ''), '미상') AS product_short,
            COALESCE(NULLIF(TRIM(i.model_name), ''), COALESCE(NULLIF(TRIM(i.product_short), ''), '미상')) AS model_name,
            COUNT(*) AS qty
        FROM inventory_items i
        WHERE i.upload_id IN ({up_ph})
          AND i.holder_type = 'partner'
        GROUP BY COALESCE(NULLIF(TRIM(i.product_short), ''), '미상'),
                 COALESCE(NULLIF(TRIM(i.model_name), ''), COALESCE(NULLIF(TRIM(i.product_short), ''), '미상'))
        ORDER BY qty DESC, product_short, model_name
        """,
        upload_ids,
    ).fetchall()
    products: dict[str, dict] = {}
    for row in rows:
        short = row["product_short"] or "미상"
        item = products.get(short)
        if not item:
            item = {"product_short": short, "qty": 0, "models": []}
            products[short] = item
        qty = int(row["qty"] or 0)
        item["qty"] += qty
        item["models"].append({"model_name": row["model_name"] or short, "qty": qty})
    catalog = sorted(products.values(), key=lambda p: (-p["qty"], p["product_short"]))
    for item in catalog:
        item["models"].sort(key=lambda m: (-m["qty"], m["model_name"]))
    return {
        "as_of_date": ",".join(as_of_dates),
        "uploads": uploads,
        "products": catalog,
    }


def inventory_model_breakdown(
    conn,
    dealer_id: str | None = None,
    region: str | None = None,
    keyword: str | None = None,
    bbox=None,
    limit: int = 20,
) -> list[dict]:
    """필터에 맞는 모든 기종 대수. 지도에 안 올린 기종도 포함한다."""
    upload_ids, _uploads = _partner_upload_filter(conn, dealer_id)
    if not upload_ids:
        return []
    extra, extra_params = _scope_filters(region, keyword, bbox)
    rows = conn.execute(
        f"""
        SELECT
            COALESCE(NULLIF(i.product_short, ''), i.model_name, '미상') AS model,
            COUNT(*) AS qty,
            COUNT(DISTINCT i.store_code) AS stores,
            SUM(CASE WHEN i.hold_days >= ? THEN 1 ELSE 0 END) AS aged_qty
        FROM inventory_items i
        LEFT JOIN stores s ON s.store_code = i.store_code
        WHERE i.upload_id IN ({','.join('?' * len(upload_ids))})
          AND i.holder_type = 'partner'
          {extra}
        GROUP BY COALESCE(NULLIF(i.product_short, ''), i.model_name, '미상')
        ORDER BY qty DESC
        LIMIT ?
        """,
        (AGED_DAYS, *upload_ids, *extra_params, limit),
    ).fetchall()
    return [
        {
            "model": r["model"],
            "qty": int(r["qty"] or 0),
            "stores": int(r["stores"] or 0),
            "aged_qty": int(r["aged_qty"] or 0),
        }
        for r in rows
    ]


def inventory_overview(conn, dealer_id: str | None = None) -> dict:
    """Gemini에 넘길 전체 재고 요약. SQL 집계만 사용한다."""
    upload_ids, uploads = _partner_upload_filter(conn, dealer_id)
    cache_key = (dealer_id or "", tuple(upload_ids))
    cached = _OVERVIEW_CACHE.get(cache_key)
    if cached is not None:
        return cached
    empty = {
        "as_of_date": "",
        "uploads": uploads,
        "total_qty": 0,
        "store_count": 0,
        "by_dealer": [],
        "by_region": [],
        "by_model": [],
        "hold_buckets": {"under_15": 0, "days_15_29": 0, "days_30_plus": 0},
        "aged_qty": 0,
        "top_aged_stores": [],
        "top_stores": [],
    }
    if not upload_ids:
        return empty
    up_ph = ",".join("?" * len(upload_ids))
    base = f"""
        FROM inventory_items i
        LEFT JOIN stores s ON s.store_code = i.store_code
        WHERE i.upload_id IN ({up_ph})
          AND i.holder_type = 'partner'
    """
    totals = conn.execute(
        f"""
        SELECT COUNT(*) AS qty,
               COUNT(DISTINCT i.store_code) AS stores,
               SUM(CASE WHEN COALESCE(i.hold_days, 0) < 15 THEN 1 ELSE 0 END) AS fresh_qty,
               SUM(CASE WHEN i.hold_days >= 15 AND i.hold_days < 30 THEN 1 ELSE 0 END) AS warn_qty,
               SUM(CASE WHEN i.hold_days >= ? THEN 1 ELSE 0 END) AS aged_qty
        {base}
        """,
        (AGED_DAYS, *upload_ids),
    ).fetchone()
    by_dealer = [
        dict(r)
        for r in conn.execute(
            f"""
            SELECT i.dealer_id, MAX(i.dealer_code) AS dealer_code, MAX(i.dealer_name) AS dealer_name,
                   COUNT(*) AS qty, COUNT(DISTINCT i.store_code) AS stores,
                   SUM(CASE WHEN i.hold_days >= ? THEN 1 ELSE 0 END) AS aged_qty
            {base}
            GROUP BY i.dealer_id
            ORDER BY qty DESC
            """,
            (AGED_DAYS, *upload_ids),
        ).fetchall()
    ]
    by_model = [
        dict(r)
        for r in conn.execute(
            f"""
            SELECT COALESCE(NULLIF(i.product_short, ''), i.model_name, '미상') AS model,
                   COUNT(*) AS qty, COUNT(DISTINCT i.store_code) AS stores,
                   SUM(CASE WHEN i.hold_days >= ? THEN 1 ELSE 0 END) AS aged_qty
            {base}
            GROUP BY COALESCE(NULLIF(i.product_short, ''), i.model_name, '미상')
            ORDER BY qty DESC
            LIMIT 12
            """,
            (AGED_DAYS, *upload_ids),
        ).fetchall()
    ]
    region_cases = []
    region_params: list = []
    for name, prefixes in REGION_PREFIXES:
        parts = []
        for prefix in prefixes:
            parts.append("REPLACE(COALESCE(s.address, ''), ' ', '') LIKE ?")
            region_params.append(prefix.replace(" ", "") + "%")
        region_cases.append(f"WHEN {' OR '.join(parts)} THEN ?")
        region_params.append(name)
    region_expr = "CASE " + " ".join(region_cases) + " ELSE '기타' END"
    by_region = [
        dict(r)
        for r in conn.execute(
            f"""
            SELECT {region_expr} AS region,
                   COUNT(*) AS qty, COUNT(DISTINCT i.store_code) AS stores,
                   SUM(CASE WHEN i.hold_days >= ? THEN 1 ELSE 0 END) AS aged_qty
            {base}
            GROUP BY region
            ORDER BY qty DESC
            """,
            (*region_params, AGED_DAYS, *upload_ids),
        ).fetchall()
    ]
    top_aged = [
        dict(r)
        for r in conn.execute(
            f"""
            SELECT i.store_code,
                   COALESCE(s.name, MAX(i.holder_name), i.store_code) AS name,
                   s.address,
                   MAX(i.dealer_name) AS dealer_name,
                   COUNT(*) AS qty,
                   SUM(CASE WHEN i.hold_days >= ? THEN 1 ELSE 0 END) AS aged_qty,
                   MAX(i.hold_days) AS max_hold_days
            {base}
            GROUP BY i.store_code
            HAVING SUM(CASE WHEN i.hold_days >= ? THEN 1 ELSE 0 END) > 0
            ORDER BY aged_qty DESC, max_hold_days DESC
            LIMIT 8
            """,
            (AGED_DAYS, *upload_ids, AGED_DAYS),
        ).fetchall()
    ]
    top_stores = [
        dict(r)
        for r in conn.execute(
            f"""
            SELECT i.store_code,
                   COALESCE(s.name, MAX(i.holder_name), i.store_code) AS name,
                   s.address,
                   COUNT(*) AS qty,
                   SUM(CASE WHEN i.hold_days >= ? THEN 1 ELSE 0 END) AS aged_qty
            {base}
            GROUP BY i.store_code
            ORDER BY qty DESC
            LIMIT 8
            """,
            (AGED_DAYS, *upload_ids),
        ).fetchall()
    ]
    as_of_dates = sorted({u.get("as_of_date") or "" for u in uploads if u.get("as_of_date")})
    aged_qty = int(totals["aged_qty"] or 0)
    uploads_by_dealer = {u.get("dealer_id") or "": u for u in uploads}
    result = {
        "as_of_date": ",".join(as_of_dates),
        "uploads": uploads,
        "total_qty": int(totals["qty"] or 0),
        "store_count": int(totals["stores"] or 0),
        "by_dealer": [
            {
                "dealer_id": d["dealer_id"] or "",
                "dealer_code": d["dealer_code"] or "",
                "dealer_name": d["dealer_name"] or "미지정",
                "qty": int(d["qty"] or 0),
                "stores": int(d["stores"] or 0),
                "aged_qty": int(d["aged_qty"] or 0),
                "filename": (uploads_by_dealer.get(d["dealer_id"] or "") or {}).get("filename") or "",
                "as_of_date": (uploads_by_dealer.get(d["dealer_id"] or "") or {}).get("as_of_date") or "",
                "uploaded_at": (uploads_by_dealer.get(d["dealer_id"] or "") or {}).get("created_at") or "",
                "has_upload": True,
            }
            for d in by_dealer
        ],
        "by_region": [
            {
                "region": r["region"],
                "qty": int(r["qty"] or 0),
                "stores": int(r["stores"] or 0),
                "aged_qty": int(r["aged_qty"] or 0),
            }
            for r in by_region
        ],
        "by_model": [
            {
                "model": m["model"],
                "qty": int(m["qty"] or 0),
                "stores": int(m["stores"] or 0),
                "aged_qty": int(m["aged_qty"] or 0),
            }
            for m in by_model
        ],
        "hold_buckets": {
            "under_15": int(totals["fresh_qty"] or 0),
            "days_15_29": int(totals["warn_qty"] or 0),
            "days_30_plus": aged_qty,
        },
        "aged_qty": aged_qty,
        "top_aged_stores": [
            {
                "store_code": r["store_code"],
                "name": r["name"],
                "address": r["address"] or "",
                "dealer_name": r["dealer_name"] or "",
                "aged_qty": int(r["aged_qty"] or 0),
                "qty": int(r["qty"] or 0),
                "max_hold_days": int(r["max_hold_days"] or 0),
            }
            for r in top_aged
        ],
        "top_stores": [
            {
                "store_code": r["store_code"],
                "name": r["name"],
                "address": r["address"] or "",
                "qty": int(r["qty"] or 0),
                "aged_qty": int(r["aged_qty"] or 0),
            }
            for r in top_stores
        ],
    }
    _OVERVIEW_CACHE[cache_key] = result
    return result


def inventory_dealer_roster(conn) -> dict:
    """관리자용 전체 대리점 목록. 포털 계정 + 재고를 올린 대리점을 모두 보여 준다."""
    upload_ids, uploads = _partner_upload_filter(conn, None)
    stats_by_id: dict[str, dict] = {}
    if upload_ids:
        up_ph = ",".join("?" * len(upload_ids))
        for row in conn.execute(
            f"""
            SELECT i.dealer_id,
                   MAX(i.dealer_code) AS dealer_code,
                   MAX(i.dealer_name) AS dealer_name,
                   COUNT(*) AS qty,
                   COUNT(DISTINCT i.store_code) AS stores,
                   SUM(CASE WHEN i.hold_days >= ? THEN 1 ELSE 0 END) AS aged_qty
            FROM inventory_items i
            WHERE i.upload_id IN ({up_ph})
              AND i.holder_type = 'partner'
            GROUP BY i.dealer_id
            """,
            (AGED_DAYS, *upload_ids),
        ):
            stats_by_id[row["dealer_id"] or ""] = dict(row)
    uploads_by_id = {u.get("dealer_id") or "": u for u in uploads}

    names: dict[str, dict] = {}
    for row in conn.execute(
        """
        SELECT DISTINCT d.id, d.dealer_code, d.name
        FROM dealers d
        JOIN admins a ON a.dealer_id = d.id
        WHERE COALESCE(a.role, '') = 'dealer'
           OR (a.dealer_id IS NOT NULL AND TRIM(a.dealer_id) != '')
        ORDER BY d.name
        """
    ):
        names[row["id"]] = {
            "dealer_id": row["id"],
            "dealer_code": row["dealer_code"] or "",
            "dealer_name": row["name"] or "",
        }
    for dealer_id, upload in uploads_by_id.items():
        if not dealer_id:
            continue
        prev = names.get(dealer_id) or {}
        names[dealer_id] = {
            "dealer_id": dealer_id,
            "dealer_code": prev.get("dealer_code") or upload.get("dealer_code") or "",
            "dealer_name": prev.get("dealer_name") or upload.get("dealer_name") or "",
        }
    for dealer_id, stats in stats_by_id.items():
        if not dealer_id or dealer_id in names:
            continue
        names[dealer_id] = {
            "dealer_id": dealer_id,
            "dealer_code": stats.get("dealer_code") or "",
            "dealer_name": stats.get("dealer_name") or "",
        }

    roster = []
    for dealer_id, meta in names.items():
        stats = stats_by_id.get(dealer_id) or {}
        upload = uploads_by_id.get(dealer_id) or {}
        roster.append(
            {
                "dealer_id": dealer_id,
                "dealer_code": meta.get("dealer_code") or stats.get("dealer_code") or upload.get("dealer_code") or "",
                "dealer_name": meta.get("dealer_name") or stats.get("dealer_name") or upload.get("dealer_name") or "미지정",
                "qty": int(stats.get("qty") or 0),
                "stores": int(stats.get("stores") or 0),
                "aged_qty": int(stats.get("aged_qty") or 0),
                "filename": upload.get("filename") or "",
                "as_of_date": upload.get("as_of_date") or "",
                "uploaded_at": upload.get("created_at") or "",
                "row_count": int(upload.get("row_count") or 0),
                "has_upload": bool(upload),
            }
        )
    roster.sort(key=lambda d: (-d["qty"], d["dealer_name"] or ""))
    uploaded = [d for d in roster if d["has_upload"]]
    return {
        "dealer_count": len(roster),
        "uploaded_count": len(uploaded),
        "pending_count": len(roster) - len(uploaded),
        "dealers": roster,
    }
